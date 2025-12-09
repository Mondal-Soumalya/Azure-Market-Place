# define "service_desk_column_process" function
def service_desk_column_process(file_path: str) -> dict[str, str]: #type: ignore
    # define missing column empty list
    missing_columns = []

    # importing python module:S01
    try:
        from pathlib import Path
        import sys
        from openpyxl import load_workbook
        import pandas
    except Exception as error:
        return {'status' : 'ERROR', 'file_name' : 'Service-Desk-Column-Process', 'step' : '1', 'message' : f'{str(error)}'}

    # appending system path:S02
    try:
        sys.path.append(str(Path.cwd()))
    except Exception as error:
        return {'status' : 'ERROR', 'file_name' : 'Service-Desk-Column-Process', 'step' : '2', 'message' : f'{str(error)}'}

    # importing "log_writer" function:S03
    try:
        from Backend.LogHandler.logwriter import log_writer
    except Exception as error:
        return {'status' : 'ERROR', 'file_name' : 'Service-Desk-Column-Process', 'step' : '3', 'message' : f'{str(error)}'}

    # define path:S04
    try:
        parent_folder_path = Path.cwd()
        backend_folder_path = Path(parent_folder_path) / 'Backend'
        service_desk_dump_handler_folder_path = Path(backend_folder_path) / 'ServiceDeskDumpHandler'
        file_process_folder_path = Path(service_desk_dump_handler_folder_path) / 'FileProcess'
        reference_folder_path = Path(file_process_folder_path) / 'ReferenceData'
        temp_files_dump_folder_path = Path(parent_folder_path) / 'TempFilesDump'
        log_writer(script_name = 'Service-Desk-Column-Process', steps = '4', status = 'SUCCESS', message = 'All Folders Are Defined')
    except Exception as error:
        log_writer(script_name = 'Service-Desk-Column-Process', steps = '4', status = 'ERROR', message = f'{str(error)}')
        return {'status' : 'ERROR', 'file_name' : 'Service-Desk-Column-Process', 'step' : '4', 'message' : f'{str(error)}'}

    # check if file is valid:S05
    try:
        # convert string file path into Path object
        file_path_object = Path(file_path)
        if file_path_object.exists() and file_path_object.is_file() and file_path_object.suffix.lower() == ".xlsx":
            log_writer(script_name = 'Service-Desk-Column-Process', steps = '5', status = 'SUCCESS', message = f'Submitted File: "{file_path_object.name}" Acceptable With Allowed File Type')
        else:
            log_writer(script_name = 'Service-Desk-Column-Process', steps = '5', status = 'ERROR', message = f'Submitted File: "{file_path_object.name}" File Type Is Not Allowed List Which Are ".csv" Or ".xlsx"')
            return {'status' : 'INFO', 'message' : f'Submitted File Not A Acceptable Format Or Type Is Not Allowed List Which Is ".xlsx"', 'missing_columns' : missing_columns} #type: ignore
    except Exception as error:
        log_writer(script_name = 'Service-Desk-Column-Process', steps = '5', status = 'ERROR', message = f'{str(error)}')
        return {'status' : 'ERROR', 'file_name' : 'Service-Desk-Column-Process', 'step' : '5', 'message' : f'{str(error)}'}

    # check workbook sheet count:S06
    try:
        # load the workbook
        ticket_workbook = load_workbook(file_path)
        # define sheet name
        ticket_sheet_names = ticket_workbook.sheetnames
        # check number of sheet
        if (int(len(ticket_sheet_names)) != 1):
            log_writer(script_name = 'Service-Desk-Column-Process', steps = '6', status = 'ERROR', message = f'Submitted File: "{file_path_object.name}" Present Multiple Sheets')
            return {'status' : 'INFO', 'message' : f'Submitted File Present Multiple Sheets', 'missing_columns' : missing_columns} #type: ignore
        else:
            sheet_name = ticket_sheet_names[0]
            log_writer(script_name = 'Service-Desk-Column-Process', steps = '6', status = 'SUCCESS', message = f'Submitted File: "{file_path_object.name}" Present One Sheet Which Is: "{sheet_name}"')
    except Exception as error:
        log_writer(script_name = 'Service-Desk-Column-Process', steps = '6', status = 'ERROR', message = f'{str(error)}')
        return {'status' : 'ERROR', 'file_name' : 'Service-Desk-Column-Process', 'step' : '6', 'message' : f'{str(error)}'}

    # load the ticket file for further processing:S07
    try:
        ticket_dataframe = pandas.read_excel(file_path_object, sheet_name = sheet_name, engine = 'openpyxl', keep_default_na = False)
        # assign "type" -> "ServiceDesk"
        ticket_dataframe['type'] = 'ServiceDesk'
        log_writer(script_name = 'Service-Desk-Column-Process', steps = '7', status = 'SUCCESS', message = f'Submitted File: "{file_path_object.name}" Sheet: "{sheet_name}" Loaded Into Memory For Processing')
    except Exception as error:
        log_writer(script_name = 'Service-Desk-Column-Process', steps = '7', status = 'ERROR', message = f'{str(error)}')
        return {'status' : 'ERROR', 'file_name' : 'Service-Desk-Column-Process', 'step' : '7', 'message' : f'{str(error)}'}

    # define column name dictonary:S08
    try:
        required_coulmn_rename_dict = {
        'ticket_number' : 'ticket_number.txt',
        'created_by' : 'created_by.txt',
        'opened_at' : 'opened_at.txt',
        'opened_by' : 'opened_by.txt',
        'contact_type' : 'contact_type.txt',
        'state' : 'state.txt',
        'priority' : 'priority.txt',
        'category' : 'category.txt',
        'sub_category' : 'sub_category.txt',
        'assignment_group' : 'assignment_group.txt',
        'resolved_by' : 'resolved_by.txt',
        'resolved_at' : 'resolved_at.txt',
        'short_description' : 'short_description.txt',
        'description' : 'description.txt',
        'work_notes' : 'work_notes.txt',
        'resolution_notes' : 'resolution_notes.txt'
        }
    except Exception as error:
        log_writer(script_name = 'Service-Desk-Column-Process', steps = '8', status = 'ERROR', message = f'{str(error)}')
        return {'status' : 'ERROR', 'file_name' : 'Service-Desk-Column-Process', 'step' : '8', 'message' : f'{str(error)}'}

    # loop through all the required column:S09
    for required_column, required_column_allowed_list_file_name in required_coulmn_rename_dict.items():
        # define column present constant
        column_present_status = False

        # load allowed list:S09-A
        try:
            allowed_list_file_path = Path(reference_folder_path) / required_column_allowed_list_file_name
            # check if file is valid
            if ((allowed_list_file_path.exists()) and (allowed_list_file_path.is_file()) and (allowed_list_file_path.suffix.lower() == '.txt')):
                with allowed_list_file_path.open('r', encoding = 'utf-8') as list_file:
                    allowed_list = [line.strip() for line in list_file if line.strip()]
                    # check count for allowed list
                    if (int(len(allowed_list)) > 0):
                        log_writer(script_name = 'Service-Desk-Column-Process', steps = '9-A', status = 'SUCCESS', message = f'Allowed List For "{required_column}" Present Inside "{required_column_allowed_list_file_name}" File')
                    else:
                        log_writer(script_name = 'Service-Desk-Column-Process', steps = '9-A', status = 'ERROR', message = f'Allowed List For "{required_column}" Not Present Inside "{required_column_allowed_list_file_name}" File')
                        return {'status' : 'ERROR', 'file_name' : 'Service-Desk-Column-Process', 'step' : '9-A', 'message' : f'No Allowed List For "{required_column}" Not Present Inside "{required_column_allowed_list_file_name}" File'}
            else:
                log_writer(script_name = 'Service-Desk-Column-Process', steps = '9-A', status = 'ERROR', message = f'File: "{required_column_allowed_list_file_name}" Not Acceptable File')
                return {'status' : 'ERROR', 'file_name' : 'Service-Desk-Column-Process', 'step' : '9-A', 'message' : f'File: "{required_column_allowed_list_file_name}" Not Acceptable File'}
        except Exception as error:
            log_writer(script_name = 'Service-Desk-Column-Process', steps = '9-A', status = 'ERROR', message = f'{str(error)}')
            return {'status' : 'ERROR', 'file_name' : 'Service-Desk-Column-Process', 'step' : '9-A', 'message' : f'{str(error)}'}

        # loop through every column for matching:S09-B
        try:
            for column_name in ticket_dataframe.columns:
                column_clean = column_name.strip().lower().replace(' ', '').replace('_', '')
                allowed_list_clean = [column_list.strip().lower().replace(' ', '').replace('_', '') for column_list in allowed_list]
                # matching column name
                if column_clean in allowed_list_clean:
                    column_present_status = True
                    ticket_dataframe.rename(columns={column_name: required_column}, inplace = True)
                    log_writer(script_name = 'Service-Desk-Column-Process', steps = '9-B', status = 'SUCCESS', message = f'Submitted File: "{file_path_object.name}" Column Name Changed To "{required_column}"')
                    break
            # if column not found inside submitted file
            if (not (column_present_status)):
                log_writer(script_name = 'Service-Desk-Column-Process', steps = '9-B', status = 'ERROR', message = f'Submitted File: "{file_path_object.name}" Column Name "{required_column}" Not Found')
                missing_columns.append(str(required_column).replace('_', ' ').title())
        except Exception as error:
            log_writer(script_name = 'Service-Desk-Column-Process', steps = '9-B', status = 'ERROR', message = f'{str(error)}')
            return {'status' : 'ERROR', 'file_name' : 'Service-Desk-Column-Process', 'step' : '9-B', 'message' : f'{str(error)}'}

    # filter only required column:S10
    try:
        required_output_column = list(required_coulmn_rename_dict.keys())
        # include "type" column explicitly in output
        required_output_column.append('type')
        # add missing columns with "N/A"
        for output_column in required_output_column:
            if output_column not in ticket_dataframe.columns:
                ticket_dataframe[output_column] = 'N/A'
        # strip whitespaces, convert empty/whitespace-only strings to "N/A"
        ticket_dataframe = ticket_dataframe.apply(lambda col: col.map(lambda x: (pandas.NA if isinstance(x, str) and x.strip() == '' else (x.strip() if isinstance(x, str) else x))))
        # fill empty cells with "N/A"
        ticket_dataframe = ticket_dataframe.fillna('N/A')
        # reorder columns to required output
        ticket_dataframe = ticket_dataframe[required_output_column]
        log_writer(script_name = 'Service-Desk-Column-Process', steps = '10', status = 'SUCCESS', message = f'Filter Output File With Required Column; Others Column Drop From The File: "{file_path_object.name}"')
    except Exception as error:
        log_writer(script_name = 'Service-Desk-Column-Process', steps = '10', status = 'ERROR', message = f'{str(error)}')
        return {'status' : 'ERROR', 'file_name' : 'Service-Desk-Column-Process', 'step' : '10', 'message' : f'{str(error)}'}

    # create updated excel file:S11
    try:
        # define output file path
        modified_output_file_path = Path(temp_files_dump_folder_path) / Path(file_path).name
        # save the file
        with pandas.ExcelWriter(modified_output_file_path, mode = 'w', engine = 'openpyxl') as excel_file_writer:
            ticket_dataframe.to_excel(excel_file_writer, sheet_name = 'raw_data', index = False)
        # check if file is saved
        if ((modified_output_file_path.exists()) and (modified_output_file_path.is_file())):
            log_writer(script_name = 'Service-Desk-Column-Process', steps = '11', status = 'SUCCESS', message = f'New File: "{modified_output_file_path.name}" Saved With Updated Column Name')
    except Exception as error:
        log_writer(script_name = 'Service-Desk-Column-Process', steps = '11', status = 'ERROR', message = f'{str(error)}')
        return {'status' : 'ERROR', 'file_name' : 'Service-Desk-Column-Process', 'step' : '11', 'message' : f'{str(error)}'}

    # define return statement:S12
    try:
        if (int(len(missing_columns)) > 0):
            return {'status' : 'INFO', 'message' : 'Required Column Not Present Inside Ticket Dump File', 'missing_columns' : missing_columns, 'file_path' : modified_output_file_path} #type: ignore
        else:
            return {'status' : 'SUCCESS', 'message' : 'Column Process Completed', 'missing_columns' : 'N/A', 'file_path' : modified_output_file_path} #type: ignore
    except Exception as error:
        log_writer(script_name = 'Service-Desk-Column-Process', steps = '12', status = 'ERROR', message = f'{str(error)}')
        return {'status' : 'ERROR', 'file_name' : 'Service-Desk-Column-Process', 'step' : '12', 'message' : f'{str(error)}'}