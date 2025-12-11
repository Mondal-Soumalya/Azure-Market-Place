def incident_file_to_db(file_unique_id: str, file_path: str) -> dict[str, str]: #type: ignore
    # importing python module:S1
    try:
        from pathlib import Path
        import sys
        from dotenv import dotenv_values
        from openpyxl import load_workbook
        import pandas
        import psycopg2
        from psycopg2.extras import execute_values
    except Exception as error:
        return {'status' : 'ERROR', 'file_name' : 'Incident-File-To-DB', 'step' : '1', 'message' : str(error)}

    # appending system path:S2
    try:
        sys.path.append(str(Path.cwd()))
    except Exception as error:
        return {'status' : 'ERROR', 'file_name' : 'Incident-File-To-DB', 'step' : '2', 'message' : str(error)}

    # importing "log_writer" function:S3
    try:
        from Backend.LogHandler.logwriter import log_writer
    except Exception as error:
        return {'status' : 'ERROR', 'file_name' : 'Incident-File-To-DB', 'step' : '3', 'message' : str(error)}

    # define folder and file path:S4
    try:
        parent_folder_path = Path.cwd()
        env_file_path = Path(parent_folder_path) / '.env'
        log_writer(script_name = 'Incident-File-To-DB', steps = '4', status = 'SUCCESS', message = 'All Folder And File Path Defined.')
    except Exception as error:
        log_writer(script_name = 'Incident-File-To-DB', steps = '4', status = 'ERROR', message = str(error))
        return {'status' : 'ERROR', 'file_name' : 'Incident-File-To-DB', 'step' : '4', 'message' : str(error)}

    # check if ".env" file is present:S5
    try:
        if ((env_file_path.exists()) and (env_file_path.is_file())):
            log_writer(script_name = 'Incident-File-To-DB', steps = '5', status = 'SUCCESS', message = '".env" File Is Present.')
        else:
            log_writer(script_name = 'Incident-File-To-DB', steps = '5', status = 'ERROR', message = '".env" File Not Present.')
            return {'status' : 'ERROR', 'file_name' : 'Incident-File-To-DB', 'step' : '5', 'message' : '".env" File Not Present.'}
    except Exception as error:
        log_writer(script_name = 'Incident-File-To-DB', steps = '5', status = 'ERROR', message = str(error))
        return {'status' : 'ERROR', 'file_name' : 'Incident-File-To-DB', 'step' : '5', 'message' : str(error)}

    # load ".env" file into script:S6
    try:
        environment_values = dotenv_values(env_file_path)
        log_writer(script_name = 'Incident-File-To-DB', steps = '6', status = 'SUCCESS', message = '".env" File Loaded Into Script.')
    except Exception as error:
        log_writer(script_name = 'Incident-File-To-DB', steps = '6', status = 'ERROR', message = str(error))
        return {'status' : 'ERROR', 'file_name' : 'Incident-File-To-DB', 'step' : '6', 'message' : str(error)}

    # define database connection parameter:S7
    try:
        database_connection_parameter = {
            "dbname" : str(environment_values.get('DATABASE_NAME')),
            "user" : str(environment_values.get('DATABASE_USER')),
            "password" : str(environment_values.get('DATABASE_PASSWORD')),
            "host" : str(environment_values.get('DATABASE_HOST')),
            "port" : str(environment_values.get('DATABASE_PORT'))
        }
        log_writer(script_name = 'Incident-File-To-DB', steps = '7', status = 'SUCCESS', message = 'Database Connection Parameter Defined.')
    except Exception as error:
        log_writer(script_name = 'Incident-File-To-DB', steps = '7', status = 'ERROR', message = str(error))
        return {'status' : 'ERROR', 'file_name' : 'Incident-File-To-DB', 'step' : '7', 'message' : str(error)}

    # check if "input_incident_data" table present inside database:S8
    try:
        input_incident_data_table_present_check_sql = '''
        SELECT EXISTS (
            SELECT FROM information_schema.tables
            WHERE table_schema = 'public'
            AND table_name = 'input_incident_data'
        );'''
        with psycopg2.connect(**database_connection_parameter) as database_connection: # type: ignore
            with database_connection.cursor() as database_cursor:
                database_cursor.execute(input_incident_data_table_present_check_sql)
                if (database_cursor.fetchone()[0]):
                    log_writer(script_name = 'Incident-File-To-DB', steps = '8', status = 'SUCCESS', message = '"input_incident_data" Table Present Inside Database.')
                else:
                    log_writer(script_name = 'Incident-File-To-DB', steps = '8', status = 'ERROR', message = '"input_incident_data" Table Not Present.')
                    return {'status' : 'ERROR', 'file_name' : 'Incident-File-To-DB', 'step' : '8', 'message' : '"input_incident_data" Table Not Present.'}
    except Exception as error:
        log_writer(script_name = 'Incident-File-To-DB', steps = '8', status = 'ERROR', message = str(error))
        return {'status' : 'ERROR', 'file_name' : 'Incident-File-To-DB', 'step' : '8', 'message' : str(error)}

    # check if "skip_row_details" table present inside database:S9
    try:
        skip_row_details_table_present_check_sql = '''
        SELECT EXISTS (
            SELECT FROM information_schema.tables
            WHERE table_schema = 'public'
            AND table_name = 'skip_row_details'
        );'''
        with psycopg2.connect(**database_connection_parameter) as database_connection: # type: ignore
            with database_connection.cursor() as database_cursor:
                database_cursor.execute(skip_row_details_table_present_check_sql)
                if (database_cursor.fetchone()[0]):
                    log_writer(script_name = 'Incident-File-To-DB', steps = '9', status = 'SUCCESS', message = '"skip_row_details" Table Present Inside Database.')
                else:
                    log_writer(script_name = 'Incident-File-To-DB', steps = '9', status = 'ERROR', message = '"skip_row_details" Table Not Present.')
                    return {'status' : 'ERROR', 'file_name' : 'Incident-File-To-DB', 'step' : '9', 'message' : '"skip_row_details" Table Not Present.'}
    except Exception as error:
        log_writer(script_name = 'Incident-File-To-DB', steps = '9', status = 'ERROR', message = str(error))
        return {'status' : 'ERROR', 'file_name' : 'Incident-File-To-DB', 'step' : '9', 'message' : str(error)}

    # check if input file is valid:S10
    try:
        file_path_object = Path(file_path)
        if ((file_path_object.exists()) and (file_path_object.is_file()) and (file_path_object.suffix.lower() == '.xlsx')):
            log_writer(script_name = 'Incident-File-To-DB', steps = '10', status = 'SUCCESS', message = f'Submitted File: "{file_path_object.name}" Acceptable With Allowed File Type.')
        else:
            log_writer(script_name = 'Incident-File-To-DB', steps = '10', status = 'ERROR', message = f'Submitted File: "{file_path_object.name}" Not A Aceeptable File.')
            return {'status' : 'ERROR', 'file_name' : 'Incident-File-To-DB', 'step' : '10', 'message' : 'Submitted File Not A Aceeptable File.'}
    except Exception as error:
        log_writer(script_name = 'Incident-File-To-DB', steps = '10', status = 'ERROR', message = str(error))
        return {'status' : 'ERROR', 'file_name' : 'Incident-File-To-DB', 'step' : '10', 'message' : str(error)}

    # check input file sheet count:S11
    try:
        incident_file_workbook_object = load_workbook(file_path)
        incident_file_sheet_names = incident_file_workbook_object.sheetnames
        if (int(len(incident_file_sheet_names)) != 1):
            log_writer(script_name = 'Incident-File-To-DB', steps = '11', status = 'ERROR', message = f'Submitted File: "{file_path_object.name}" Present Multiple Sheets.')
            return {'status' : 'ERROR', 'file_name' : 'Incident-File-To-DB', 'step' : '11', 'message' : 'Submitted File Present Multiple Sheets.'}
        else:
            incident_file_sheet_name = incident_file_sheet_names[0]
            log_writer(script_name = 'Incident-File-To-DB', steps = '11', status = 'SUCCESS', message = f'Submitted File: "{file_path_object.name}" Present One Sheet Which Is: "{incident_file_sheet_name}".')
    except Exception as error:
        log_writer(script_name = 'Incident-File-To-DB', steps = '11', status = 'ERROR', message = str(error))
        return {'status' : 'ERROR', 'file_name' : 'Incident-File-To-DB', 'step' : '11', 'message' : str(error)}

    # load the input file for further processing:S12
    try:
        incident_ticket_dataframe = pandas.read_excel(str(file_path_object), sheet_name = incident_file_sheet_name, engine = 'openpyxl', keep_default_na = False)
        log_writer(script_name = 'Incident-File-To-DB', steps = '12', status = 'SUCCESS', message = f'Submitted File: "{file_path_object.name}" Sheet: "{incident_file_sheet_name}" Loaded Into Memory For Processing.')
    except Exception as error:
        log_writer(script_name = 'Incident-File-To-DB', steps = '12', status = 'ERROR', message = str(error))
        return {'status': 'ERROR', 'file_name': 'Incident-File-To-DB', 'step': '12', 'message': str(error)}

    # define "input_incident_data" upsert SQL:S13
    try:
        input_incident_data_insert_sql = '''
        INSERT INTO input_incident_data (
            ticket_number,
            ticket_type,
            opened_at,
            cmdb_ci,
            state,
            priority,
            category,
            channel,
            assignment_group,
            parent_ticket,
            assigned_to,
            resolved_by,
            resolved_at,
            short_description,
            description,
            work_notes,
            resolution_notes
        )
        VALUES %s
        ON CONFLICT (ticket_number)
        DO UPDATE SET
            ticket_type         = EXCLUDED.ticket_type,
            opened_at           = EXCLUDED.opened_at,
            cmdb_ci             = EXCLUDED.cmdb_ci,
            state               = EXCLUDED.state,
            priority            = EXCLUDED.priority,
            category            = EXCLUDED.category,
            channel             = EXCLUDED.channel,
            assignment_group    = EXCLUDED.assignment_group,
            parent_ticket       = EXCLUDED.parent_ticket,
            assigned_to         = EXCLUDED.assigned_to,
            resolved_by         = EXCLUDED.resolved_by,
            resolved_at         = EXCLUDED.resolved_at,
            short_description   = EXCLUDED.short_description,
            description         = EXCLUDED.description,
            work_notes          = EXCLUDED.work_notes,
            resolution_notes    = EXCLUDED.resolution_notes,
            row_updated_at      = NOW(),
            row_status          = 1;'''
        log_writer(script_name = 'Incident-File-To-DB', steps = '13', status = 'SUCCESS', message = 'Upsert SQL Defined For "input_incident_data" Table.')
    except Exception as error:
        log_writer(script_name = 'Incident-File-To-DB', steps = '13', status = 'ERROR', message = str(error))
        return {'status': 'ERROR', 'file_name': 'Incident-File-To-DB', 'step': '13', 'message': str(error)}

    # define "skip_row_details" insert sql:S14
    try:
        skip_row_details_insert_sql = '''
        INSERT INTO skip_row_details (
            file_unique_id,
            ticket_number,
            skip_reason
        )
        VALUES %s;'''
        log_writer(script_name = 'Incident-File-To-DB', steps = '14', status = 'SUCCESS', message = 'Insert SQL Defined For "skip_row_details" Table.')
    except Exception as error:
        log_writer(script_name = 'Incident-File-To-DB', steps = '14', status = 'ERROR', message = str(error))
        return {'status': 'ERROR', 'file_name': 'Incident-File-To-DB', 'step': '14', 'message': str(error)}

    # define "normalize_datetime" function:S15
    try:
        def normalize_datetime(value):
            if ((pandas.isna(value)) or (value == '')):
                return None
            try:
                if isinstance(value, (int, float)):
                    dt = pandas.to_datetime(value, unit = 'd', origin = '1899-12-30', errors = 'coerce')
                else:
                    dt = pandas.to_datetime(value, errors = 'coerce', dayfirst = True)
            except Exception:
                return None
            if (pandas.isna(dt)):
                return None
            if dt.tzinfo is None:
                return dt.tz_localize('UTC')
            return dt.tz_convert('UTC')
    except Exception as error:
        log_writer(script_name = 'Incident-File-To-DB', steps = '15', status = 'ERROR', message = str(error))
        return {'status': 'ERROR', 'file_name': 'Incident-File-To-DB', 'step': '15', 'message': str(error)}

    # define constant
    file_to_db_rows_limiter = int(str(environment_values.get('FILE_TO_DB_BATCH')))
    input_data_insert_rows = []
    skip_data_insert_rows = []
    row_counter = 0
    skipped_rows = 0
    required_check_columns = ['ticket_number', 'opened_at']

    # loop through all the rows
    for _, row in incident_ticket_dataframe.iterrows():
        row_counter += 1

        # skip invalid rows
        if any(pandas.isna(row.get(col)) or str(row.get(col)).strip() in ['', 'N/A'] for col in required_check_columns):
            skipped_rows += 1
            if not (pandas.isna(row.get('ticket_number')) or str(row.get('ticket_number')).strip() in ['', 'N/A']):
                if pandas.isna(row.get('opened_at')) or str(row.get('opened_at')).strip() in ['', 'N/A']:
                    skip_data_insert_rows.append((str(file_unique_id), str(row.get('ticket_number')), 'Open Date'))
            continue

        # normalize datetime for upserting:S16
        try:
            normalize_opened_at = normalize_datetime(row.get('opened_at'))
            raw_resolved_at = row.get('resolved_at')
            normalize_resolved_at = normalize_datetime(raw_resolved_at)
            # if "resolved_at" missing → use "opened_at" & log skip
            if normalize_resolved_at is None:
                normalize_resolved_at = normalize_opened_at
                skip_data_insert_rows.append((str(file_unique_id), str(row.get('ticket_number')), 'Resolved Date'))

            if normalize_opened_at is None:
                skipped_rows += 1
                skip_data_insert_rows.append((str(file_unique_id), str(row.get('ticket_number')), 'Open Date'))
                continue
        except Exception as error:
            log_writer(script_name='Incident-File-To-DB', steps='16', status='ERROR', message=str(error))
            return {'status': 'ERROR', 'file_name': 'Incident-File-To-DB', 'step': '16', 'message': str(error)}

        # prepare value tuple for upserting:S17
        try:
            def safe_get(row, field):
                value = row.get(field)
                if ((pandas.isna(value)) or (str(value).strip() == '')):
                    return 'N/A'
                return value

            input_data_insert_rows.append((
                str(row.get('ticket_number')),
                safe_get(row, 'type'),
                normalize_opened_at,
                safe_get(row, 'cmdb_ci'),
                safe_get(row, 'state'),
                safe_get(row, 'priority'),
                safe_get(row, 'category'),
                safe_get(row, 'channel'),
                safe_get(row, 'assignment_group'),
                safe_get(row, 'parent'),
                safe_get(row, 'assigned_to'),
                safe_get(row, 'resolved_by'),
                normalize_resolved_at,
                safe_get(row, 'short_description'),
                safe_get(row, 'description'),
                safe_get(row, 'work_notes'),
                safe_get(row, 'resolution_notes')
            ))

        except Exception as error:
            log_writer(script_name = 'Incident-File-To-DB', steps = '17', status = 'ERROR', message = str(error))
            return {'status': 'ERROR', 'file_name': 'Incident-File-To-DB', 'step': '17', 'message': str(error)}

        # upsert input batch when limit reached:S18
        if (len(input_data_insert_rows) >= file_to_db_rows_limiter):
            try:
                with psycopg2.connect(**database_connection_parameter) as database_connection: #type: ignore
                    with database_connection.cursor() as database_cursor:
                        execute_values(database_cursor, input_incident_data_insert_sql, input_data_insert_rows)
                        database_connection.commit()
                        log_writer(script_name = 'Incident-File-To-DB', steps = '18', status = 'SUCCESS', message = f'File: "{file_path_object.name}" Batch {row_counter - len(input_data_insert_rows) + 1}-{row_counter} Insert Success In "input_incident_data" Table.')
                        input_data_insert_rows.clear()
            except Exception as error:
                database_connection.rollback()
                log_writer(script_name = 'Incident-File-To-DB', steps = '18', status = 'ERROR', message = str(error))
                return {'status': 'ERROR', 'file_name': 'Incident-File-To-DB', 'step': '18', 'message': str(error)}

        # upsert skip batch when limit reached:S19
        if (len(skip_data_insert_rows) >= file_to_db_rows_limiter):
            try:
                with psycopg2.connect(**database_connection_parameter) as database_connection: #type: ignore
                    with database_connection.cursor() as database_cursor:
                        execute_values(database_cursor, skip_row_details_insert_sql, skip_data_insert_rows)
                        database_connection.commit()
                        log_writer(script_name = 'Incident-File-To-DB', steps = '19', status = 'SUCCESS', message = f'File: "{file_path_object.name}" Batch {row_counter - len(skip_data_insert_rows) + 1}-{row_counter} Insert Success Inside "skip_row_details" Table.')
                        skip_data_insert_rows.clear()
            except Exception as error:
                database_connection.rollback()
                log_writer(script_name = 'Incident-File-To-DB', steps = '19', status = 'ERROR', message = str(error))
                return {'status': 'ERROR', 'file_name': 'Incident-File-To-DB', 'step': '19', 'message': str(error)}

    # handle leftover input batch rows:S20
    if (input_data_insert_rows):
        try:
            with psycopg2.connect(**database_connection_parameter) as database_connection: #type: ignore
                with database_connection.cursor() as database_cursor:
                    execute_values(database_cursor, input_incident_data_insert_sql, input_data_insert_rows)
                    database_connection.commit()
                    log_writer(script_name = 'Incident-File-To-DB', steps = '20', status = 'SUCCESS', message = f'File: "{file_path_object.name}" Batch {row_counter - len(input_data_insert_rows) + 1}-{row_counter} Insert Success Inside "input_incident_data" Table.')
                    log_writer(script_name = 'Incident-File-To-DB', steps = '20', status = 'INFO', message = f'File: "{file_path_object.name}" Total Skipped Rows: "{skipped_rows}".')
                    input_data_insert_rows.clear()
        except Exception as error:
            database_connection.rollback()
            log_writer(script_name = 'Incident-File-To-DB', steps = '20', status = 'ERROR', message=str(error))
            return {'status': 'ERROR', 'file_name': 'Incident-File-To-DB', 'step': '20', 'message': str(error)}

    # handle leftover skip batch rows:S21
    if (skip_data_insert_rows):
        try:
            with psycopg2.connect(**database_connection_parameter) as database_connection: #type: ignore
                with database_connection.cursor() as database_cursor:
                    execute_values(database_cursor, skip_row_details_insert_sql, skip_data_insert_rows)
                    database_connection.commit()
                    log_writer(script_name = 'Incident-File-To-DB', steps = '21', status = 'SUCCESS', message = f'File: "{file_path_object.name}" Batch {row_counter - len(skip_data_insert_rows) + 1}-{row_counter} Insert Success Inside "skip_row_details" Table.')
                    skip_data_insert_rows.clear()
        except Exception as error:
            database_connection.rollback()
            log_writer(script_name = 'Incident-File-To-DB', steps = '21', status = 'ERROR', message=str(error))
            return {'status': 'ERROR', 'file_name': 'Incident-File-To-DB', 'step': '21', 'message': str(error)}

    # delete temp stored file from the server:S22
    try:
        if ((file_path_object.exists()) and (file_path_object.is_file())):
            file_path_object.unlink()
            if (not file_path_object.exists()):
                log_writer(script_name = 'Incident-File-To-DB', steps = '22', status = 'SUCCESS', message = f'Temp Stored File: "{file_path_object.name}" Deleted From The "../TempFilesDump" Folder.')
                return {'status': 'SUCCESS', 'file_name': 'Incident-File-To-DB', 'step': '22', 'message': 'File To Database Data Transfer Complete.'}
            else:
                log_writer(script_name = 'Incident-File-To-DB', steps = '22', status = 'ERROR', message = f'Temp Stored File: "{file_path_object.name}" Not Deleted From The "../TempFilesDump" Folder')
                return {'status': 'ERROR', 'file_name': 'Incident-File-To-DB', 'step': '22', 'message': f'Temp Stored File: "{file_path_object.name}" Not Deleted From The "../TempFilesDump" Folder'}
        else:
            log_writer(script_name = 'Incident-File-To-DB', steps = '22', status = 'ERROR', message = f'Temp Stored File: "{file_path_object.name}" Not Present Or Not A Valid File.')
            return {'status': 'ERROR', 'file_name': 'Incident-File-To-DB', 'step': '22', 'message': f'Temp Stored File: "{file_path_object.name}" Not Present Or Not A Valid File.'}
    except Exception as error:
        log_writer(script_name = 'Incident-File-To-DB', steps = '22', status = 'ERROR', message=str(error))
        return {'status': 'ERROR', 'file_name': 'Incident-File-To-DB', 'step': '22', 'message': str(error)}